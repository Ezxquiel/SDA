from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from datetime import datetime
import os

class Colors:
    PRIMARY = "2980B9"
    LIGHT_BLUE = "D6EAF8"
    WHITE = "FFFFFF"
    HEADER_BLUE = "3498DB"

class AttendanceExcelReport:
    def __init__(self, resumen, totales, fecha_inicio, fecha_fin):
        self.resumen = resumen
        self.totales = totales
        self.fecha_inicio = fecha_inicio
        self.fecha_fin = fecha_fin
        self.wb = Workbook()
        self.ws = self.wb.active
        self.ws.title = "Reporte de Asistencia"
        self.setup_styles()

    def setup_styles(self):
        self.header_style = {
            'font': Font(bold=True, color="FFFFFF"),
            'fill': PatternFill(start_color=Colors.HEADER_BLUE, end_color=Colors.HEADER_BLUE, fill_type="solid"),
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }
        
        self.cell_style = {
            'alignment': Alignment(horizontal='center', vertical='center'),
            'border': Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        }

    def apply_style(self, cell, style_dict):
        for key, value in style_dict.items():
            setattr(cell, key, value)

    def add_totales(self):
        self.ws.append([])
        self.ws.append(['Resumen General'])
        start_row = self.ws.max_row
        
        # Aplicar estilo al título
        cell = self.ws.cell(row=start_row, column=1)
        cell.font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        if self.totales:
            self.ws.append(['Total Asistidos', self.totales["total_asistidos"]])
            self.ws.append(['Total Masculino', self.totales["total_masculino"]])
            self.ws.append(['Total Femenino', self.totales["total_femenino"]])
            self.ws.append(['Total Inasistidos', self.totales["total_inasistidos"]])
            
            total_alumnos = self.totales["total_inasistidos"] + self.totales["total_asistidos"]
            self.ws.append(['Total de Alumnos', total_alumnos])
            self.ws.append(['% Asistencia', f"{self.totales['porcentaje_asistencia']}%"])

    def add_detail_table(self):
        self.ws.append([])
        self.ws.append(['Detalle por Sección'])
        title_row = self.ws.max_row
        self.ws.cell(row=title_row, column=1).font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        headers = ['Fecha', 'Año', 'Sección', 'Asist.', 'M', 'F', 'Inasist.', 'Total de alumnos', '% Asist.']
        self.ws.append(headers)
        
        # Aplicar estilos a los encabezados
        header_row = self.ws.max_row
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col)
            self.apply_style(cell, self.header_style)
        
        for row in self.resumen:
            total = row['total_inasistidos'] + row['total_asistidos']
            row_data = [
                row['fecha_entrada'],
                row['año'],
                row['seccion'],
                row['total_asistidos'],
                row['total_masculino'],
                row['total_femenino'],
                row['total_inasistidos'],
                total,
                f"{row['porcentaje_asistencia']}%"
            ]
            self.ws.append(row_data)
            
            # Aplicar estilos a las celdas
            current_row = self.ws.max_row
            for col in range(1, len(row_data) + 1):
                cell = self.ws.cell(row=current_row, column=col)
                self.apply_style(cell, self.cell_style)
                
                # Alternar colores de fondo
                if current_row % 2 == 0:
                    cell.fill = PatternFill(start_color=Colors.LIGHT_BLUE, 
                                          end_color=Colors.LIGHT_BLUE, 
                                          fill_type="solid")

    def add_absence_codes_table(self):
        self.ws.append([])
        self.ws.append(['Detalle de Códigos de Inasistencia'])
        title_row = self.ws.max_row
        self.ws.cell(row=title_row, column=1).font = Font(bold=True, size=14, color=Colors.PRIMARY)
        
        headers = ['Año', 'Sección', 'Inasistencias']
        self.ws.append(headers)
        
        # Aplicar estilos a los encabezados
        header_row = self.ws.max_row
        for col, header in enumerate(headers, 1):
            cell = self.ws.cell(row=header_row, column=col)
            self.apply_style(cell, self.header_style)
        
        for row in self.resumen:
            if row['codigos_inasistidos'] and row['codigos_inasistidos'] != "[]":
                codigos = self.format_absence_codes(row['codigos_inasistidos'])
                self.ws.append([row['año'], row['seccion'], codigos])
                
                # Aplicar estilos a las celdas
                current_row = self.ws.max_row
                for col in range(1, 4):
                    cell = self.ws.cell(row=current_row, column=col)
                    self.apply_style(cell, self.cell_style)

    def format_absence_codes(self, codes_str):
        if not codes_str or codes_str == "[]":
            return "-"
        try:
            codes = eval(codes_str)
            code_count = {}
            for code in codes:
                code_count[code] = code_count.get(code, 0) + 1
            return ", ".join([f"{code}({count})" for code, count in code_count.items()])
        except:
            return codes_str

    def generate(self):
        try:
            self.add_totales()
            self.add_detail_table()
            self.add_absence_codes_table()
            
            # Ajustar anchos de columna
            for column in self.ws.columns:
                max_length = 0
                column = list(column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                self.ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            desktop_path = os.path.expanduser("~/Desktop")
            filename = os.path.join(desktop_path, 
                                  f'reporte_asistencia_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
            self.wb.save(filename)
            return filename
        except Exception as e:
            raise Exception(f"Error al generar el Excel: {str(e)}")